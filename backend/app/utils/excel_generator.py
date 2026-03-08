import os
import tempfile
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

def render_excel_from_invoice(invoice, services, total, provider, customization=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuenta de Cobro"
    
    # --- AESTHETIC REFINEMENT: Hide Gridlines ---
    ws.sheet_view.showGridLines = False

    # Color Constants
    HEADER_GREEN = "1B4332"  # Matches its PDF green #1b4332
    WHITE = "FFFFFF"
    BORDER_COLOR = "CCCCCC"

    # Define Styles
    header_font = Font(name='Open Sans', size=14, bold=True, color=WHITE)
    label_font = Font(name='Open Sans', size=10, bold=True, color=HEADER_GREEN)
    data_font = Font(name='Open Sans', size=10)
    table_header_font = Font(name='Open Sans', size=10, bold=True, color=WHITE)
    table_font = Font(name='Open Sans', size=10)
    
    header_fill = PatternFill(start_color=HEADER_GREEN, end_color=HEADER_GREEN, fill_type="solid")
    
    thin_side = Side(style='thin', color=BORDER_COLOR)
    header_side = Side(style='thin', color=HEADER_GREEN)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)

    # Column Widths
    ws.column_dimensions['A'].width = 14  # Fecha
    ws.column_dimensions['B'].width = 48  # Servicio/Material (slighly wider)
    ws.column_dimensions['C'].width = 12  # Cantidad
    ws.column_dimensions['D'].width = 18  # V. Unitario
    ws.column_dimensions['E'].width = 18  # V. Parcial

    # --- 1. HEADER AREA ---
    # Logo placement (rows 1-4, col A)
    if customization and customization.logo_url and customization.include_logo:
        try:
            # We add it slightly smaller to avoid overlapping
            response = requests.get(customization.logo_url, timeout=5)
            if response.status_code == 200:
                img_data = BytesIO(response.content)
                img = XLImage(img_data)
                img.width = 120
                img.height = 70
                ws.add_image(img, 'A1')
        except Exception:
            pass

    # Title "CUENTA DE COBRO N° XXX" (centered across B-E)
    ws.merge_cells('B2:E2')
    title_cell = ws['B2']
    title_cell.value = f"CUENTA DE COBRO N° {invoice.invoice_number}"
    title_cell.font = Font(name='Open Sans', size=18, bold=True, color=HEADER_GREEN)
    title_cell.alignment = Alignment(horizontal='center')

    # Date
    ws.merge_cells('B3:E3')
    date_cell = ws['B3']
    invoice_date = invoice.invoice_date.strftime('%d/%m/%Y') if hasattr(invoice.invoice_date, 'strftime') else str(invoice.invoice_date)
    date_cell.value = f"Fecha: {invoice_date}"
    date_cell.font = Font(name='Open Sans', size=11, color="666666")
    date_cell.alignment = Alignment(horizontal='center')

    # Spacer
    row = 5

    # --- 2. INFO BOX ---
    # Using a border for the info area
    info_start_row = row
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].value = "INFORMACIÓN GENERAL"
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = PatternFill(start_color="F1F6F4", end_color="F1F6F4", fill_type="solid")
    row += 1

    client_name = invoice.service_account.client.name
    ws[f'A{row}'].value = "CLIENTE:"
    ws[f'A{row}'].font = label_font
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].value = client_name
    ws[f'B{row}'].font = data_font
    row += 1

    start_date = invoice.service_account.start_date.strftime('%d/%m/%Y') if hasattr(invoice.service_account.start_date, 'strftime') else str(invoice.service_account.start_date)
    end_date = invoice.service_account.end_date.strftime('%d/%m/%Y') if hasattr(invoice.service_account.end_date, 'strftime') else str(invoice.service_account.end_date)
    ws[f'A{row}'].value = "PERIODO:"
    ws[f'A{row}'].font = label_font
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].value = f"{start_date} — {end_date}"
    ws[f'B{row}'].font = data_font
    row += 1

    if customization and customization.service_text:
        ws[f'A{row}'].value = "SERVICIO:"
        ws[f'A{row}'].font = label_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].value = customization.service_text
        ws[f'B{row}'].font = data_font
        row += 1

    row += 2 # Spacer

    # --- 3. SERVICES TABLE ---
    # Table Header
    ws.row_dimensions[row].height = 25 # Increased header height
    table_headers = ["FECHA", "SERVICIO / MATERIAL", "CANTIDAD", "V. UNITARIO", "V. PARCIAL"]
    for col_idx, text in enumerate(table_headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = text
        cell.font = table_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    row += 1

    # Table Body
    # Zebra stripes color
    ZEBRA_FILL = PatternFill(start_color="F9FAF9", end_color="F9FAF9", fill_type="solid")
    
    for i, s in enumerate(services):
        ws.row_dimensions[row].height = 20 # Increased row height
        
        s_date = s.service_date.strftime('%d/%m/%Y') if hasattr(s.service_date, 'strftime') else str(s.service_date)
        
        if s.custom_material:
            s_name = f"Viaje de {s.custom_material}"
        elif s.material:
            s_name = f"Viaje de {s.material.name}"
        else:
            s_name = "Viaje sin material"
 
        data = [s_date, s_name, s.quantity, s.price, s.total_amount]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = table_font
            cell.border = border
            
            # Apply Zebra Stripe
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
                
            if col_idx in [1, 3]: # Date, Quantity
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [4, 5]: # Prices
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '"$"#,##0'
            else: # Name
                cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    # --- 4. TOTALS AND PAYMENT SECTION ---
    row += 1
    # Total a pagar box
    ws.merge_cells(f'C{row}:D{row+1}')
    total_label = ws[f'C{row}']
    total_label.value = "TOTAL A PAGAR:"
    total_label.font = Font(name='Open Sans', size=12, bold=True, color=HEADER_GREEN)
    total_label.alignment = Alignment(horizontal='right', vertical='center')

    # Apply border to ALL cells in the merged range to ensure it's complete
    thick_border = Border(left=Side(style='medium', color=HEADER_GREEN), 
                          right=Side(style='medium', color=HEADER_GREEN),
                          top=Side(style='medium', color=HEADER_GREEN),
                          bottom=Side(style='medium', color=HEADER_GREEN))
    
    ws.merge_cells(f'E{row}:E{row+1}')
    for r in range(row, row + 2):
        ws.cell(row=r, column=5).border = thick_border
        
    total_val = ws[f'E{row}']
    total_val.value = total
    total_val.font = Font(name='Open Sans', size=16, bold=True, color=HEADER_GREEN)
    total_val.fill = PatternFill(fill_type=None) 
    total_val.alignment = Alignment(horizontal='center', vertical='center')
    total_val.number_format = '"$"#,##0'
    
    row += 1 # Reduced spacing for a more compact look

    # --- 5. SIGNATURE ---
    # Centered signature placement
    sig_row = row + 3 # 3 rows below the total for clear space
    
    if customization and customization.signature_url and customization.include_signature:
        try:
            response = requests.get(customization.signature_url, timeout=5)
            if response.status_code == 200:
                sig_data = BytesIO(response.content)
                sig_img = XLImage(sig_data)
                sig_img.width = 160
                sig_img.height = 90
                # Placing it anchored in D to roughly center it over the C-E area
                ws.add_image(sig_img, f'D{sig_row}')
                # We move sig_row 5 rows down to ensure there is no overlap with the name
                sig_row += 5
        except Exception:
            # If no image, we still need space
            sig_row += 2
            pass

    # No line as requested. Just name and document.
    ws.merge_cells(f'C{sig_row}:E{sig_row}')
    name_cell = ws[f'C{sig_row}']
    name_cell.value = provider.name
    name_cell.font = Font(name='Open Sans', size=14, bold=True, italic=True, color=HEADER_GREEN)
    name_cell.alignment = Alignment(horizontal='center', vertical='center')
    sig_row += 1

    if provider.document_number:
        ws.merge_cells(f'C{sig_row}:E{sig_row}')
        doc_cell = ws[f'C{sig_row}']
        doc_cell.value = f"CC: {provider.document_number}"
        doc_cell.font = Font(name='Open Sans', size=11, color="666666")
        doc_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Footer message
    if customization and customization.include_footer and customization.footer_message:
        row = max(row, sig_row) + 2
        ws.merge_cells(f'A{row}:E{row}')
        footer_cell = ws[f'A{row}']
        footer_cell.value = customization.footer_message
        footer_cell.font = Font(name='Open Sans', size=10, italic=True, color=HEADER_GREEN)
        footer_cell.alignment = Alignment(horizontal='center')

    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
